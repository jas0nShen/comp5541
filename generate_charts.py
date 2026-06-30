#!/usr/bin/env python3
"""
Generate three publication-ready visualization charts for LLM benchmark evaluation.
Uses openpyxl directly (no pandas) to avoid numpy/pandas version conflicts.
"""

import openpyxl
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from collections import Counter
import warnings
warnings.filterwarnings('ignore')

# Configuration
FILE = 'records_final_with_scores.xlsx'
OUTPUT_DIR = '.'

# Professional academic color palette
COLORS = {
    'GPT':      '#E07A3A',   # Warm burnt orange
    'DeepSeek': '#2CA6A4',   # Tech teal
    'Claude':   '#1B3A5C',   # Deep navy
}
BG_COLOR = '#FAFBFC'
GRID_COLOR = '#E8ECF0'

# Load Data with openpyxl
wb = openpyxl.load_workbook(FILE, data_only=True)
print('Sheet names:', wb.sheetnames)

def extract_scores(sheet_name):
    """Extract gpt_score, deepseek_score, claude_score from a sheet."""
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    print(f'\n{sheet_name} headers: {headers}')
    
    # Find score column indices
    col_map = {}
    for i, h in enumerate(headers):
        if h and isinstance(h, str):
            h_lower = h.strip().lower()
            if 'gpt_score' in h_lower:
                col_map['GPT'] = i
            elif 'deepseek_score' in h_lower:
                col_map['DeepSeek'] = i
            elif 'claude_score' in h_lower:
                col_map['Claude'] = i
    
    print(f'  Score columns found: {col_map}')
    
    scores = {'GPT': [], 'DeepSeek': [], 'Claude': []}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        for model, ci in col_map.items():
            val = row[ci]
            if val is not None:
                try:
                    scores[model].append(float(val))
                except (ValueError, TypeError):
                    pass
    
    for m, s in scores.items():
        print(f'  {m}: {len(s)} scores, sample: {s[:5]}')
    return scores

scores_reasoning = extract_scores('Sheet1')
scores_semantic  = extract_scores('Sheet2')

# Combine all scores
scores_all = {}
for m in ['GPT', 'DeepSeek', 'Claude']:
    scores_all[m] = scores_reasoning[m] + scores_semantic[m]

# Compute Statistics
avg_scores = {}
for m in ['GPT', 'DeepSeek', 'Claude']:
    avg_scores[m] = {
        'Reasoning Ability':      np.mean(scores_reasoning[m]) if scores_reasoning[m] else 0,
        'Semantic Understanding': np.mean(scores_semantic[m]) if scores_semantic[m] else 0,
    }

print('\nAverage Scores:')
for m, s in avg_scores.items():
    print(f'  {m}: {s}')

# Score distributions
score_dist = {}
for m in ['GPT', 'DeepSeek', 'Claude']:
    score_dist[m] = Counter(int(x) for x in scores_all[m])
    print(f'\n{m} distribution: {dict(sorted(score_dist[m].items(), reverse=True))}')

# Global Style Setup
sns.set_style('whitegrid')
plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Helvetica Neue', 'Helvetica', 'Arial', 'DejaVu Sans'],
    'font.size': 11,
    'axes.titlesize': 15,
    'axes.labelsize': 12,
    'xtick.labelsize': 11,
    'ytick.labelsize': 11,
    'legend.fontsize': 10.5,
    'figure.dpi': 300,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'savefig.pad_inches': 0.3,
    'axes.edgecolor': '#CCCCCC',
    'axes.linewidth': 0.8,
    'grid.color': GRID_COLOR,
    'grid.linewidth': 0.5,
})

models = ['GPT', 'DeepSeek', 'Claude']
categories = ['Reasoning Ability', 'Semantic Understanding']


# Chart 1: Grouped Bar Chart – Overall Performance Comparison
def create_chart1():
    fig, ax = plt.subplots(figsize=(8, 5.5))
    fig.patch.set_facecolor(BG_COLOR)
    ax.set_facecolor(BG_COLOR)

    x = np.arange(len(categories))
    bar_width = 0.22
    offsets = [-bar_width, 0, bar_width]

    for i, model in enumerate(models):
        vals = [avg_scores[model][cat] for cat in categories]
        bars = ax.bar(
            x + offsets[i], vals, bar_width,
            label=model, color=COLORS[model],
            edgecolor='white', linewidth=0.8,
            zorder=3, alpha=0.92
        )
        # Data labels on top
        for bar, val in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.06,
                f'{val:.2f}', ha='center', va='bottom',
                fontsize=10.5, fontweight='bold', color=COLORS[model]
            )

    ax.set_ylabel('Average Score', fontweight='semibold', labelpad=10)
    ax.set_xticks(x)
    ax.set_xticklabels(categories, fontweight='medium')
    ax.set_ylim(0, 5.6)
    ax.set_yticks([0, 1, 2, 3, 4, 5])

    # Reference line at max score
    ax.axhline(y=5, color='#B0B0B0', linestyle='--', linewidth=0.7, alpha=0.6, zorder=1)

    ax.legend(
        loc='upper left', frameon=True, framealpha=0.95,
        edgecolor='#DDDDDD', fancybox=True
    )
    ax.set_title(
        'Overall Performance Comparison Across Evaluation Dimensions',
        fontweight='bold', pad=16, fontsize=14
    )
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    out = f'{OUTPUT_DIR}/chart1_grouped_bar.png'
    fig.savefig(out)
    print(f'\n✅ Saved: {out}')
    plt.close()


# Chart 2: Donut Charts – Score Distribution
def create_chart2():
    fig, axes = plt.subplots(1, 3, figsize=(13, 4.8))
    fig.patch.set_facecolor(BG_COLOR)

    # Tier colors (harmonious, muted tones)
    tier_colors = {
        5: '#2ECC71',   # Emerald green
        4: '#5DADE2',   # Sky blue
        3: '#F4D03F',   # Soft gold
        2: '#E67E22',   # Carrot orange
        1: '#E74C3C',   # Soft red
    }
    score_labels_map = {5: 'Score 5', 4: 'Score 4', 3: 'Score 3', 2: 'Score 2', 1: 'Score 1'}

    # Determine all unique scores across all models
    all_scores_set = set()
    for m in models:
        all_scores_set.update(score_dist[m].keys())
    all_scores_sorted = sorted(all_scores_set, reverse=True)

    for idx, model in enumerate(models):
        ax = axes[idx]
        ax.set_facecolor(BG_COLOR)

        dist = score_dist[model]
        scores_present = [s for s in all_scores_sorted if dist.get(s, 0) > 0]
        sizes = [dist[s] for s in scores_present]
        colors = [tier_colors.get(s, '#AAAAAA') for s in scores_present]
        total = sum(sizes)

        wedges, texts, autotexts = ax.pie(
            sizes, labels=None,
            autopct=lambda p: f'{p:.1f}%' if p > 3 else '',
            startangle=90, colors=colors,
            pctdistance=0.78,
            wedgeprops={'width': 0.42, 'edgecolor': 'white', 'linewidth': 2.5},
            textprops={'fontsize': 9, 'fontweight': 'bold', 'color': '#333'}
        )
        for at in autotexts:
            at.set_fontsize(9)
            at.set_fontweight('bold')
            at.set_color('#333333')

        # Center text
        ax.text(0, 0.06, model, ha='center', va='center',
                fontsize=14, fontweight='bold', color=COLORS[model])
        ax.text(0, -0.18, f'n = {total}', ha='center', va='center',
                fontsize=9, color='#888888')

    # Shared legend at bottom
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor=tier_colors[s], edgecolor='white', label=score_labels_map[s])
        for s in all_scores_sorted if s in tier_colors
    ]
    fig.legend(
        handles=legend_elements, loc='lower center',
        ncol=len(legend_elements), frameon=True, framealpha=0.95,
        edgecolor='#DDDDDD', fontsize=10, fancybox=True,
        bbox_to_anchor=(0.5, -0.02)
    )

    fig.suptitle(
        'Score Distribution by Model (Combined Benchmark)',
        fontsize=14, fontweight='bold', y=1.02
    )
    plt.tight_layout()
    out = f'{OUTPUT_DIR}/chart2_donut_distribution.png'
    fig.savefig(out, bbox_inches='tight')
    print(f'✅ Saved: {out}')
    plt.close()


# Chart 3: Radar Chart – Model Capability Spectrum
def create_chart3():
    dim_labels = []
    dims = []

    # 1. Reasoning Ability (avg)
    dim_labels.append('Reasoning\nAbility')
    dims.append({m: avg_scores[m]['Reasoning Ability'] for m in models})

    # 2. Semantic Understanding (avg)
    dim_labels.append('Semantic\nUnderstanding')
    dims.append({m: avg_scores[m]['Semantic Understanding'] for m in models})

    # 3. Consistency (% of Score 5, scaled to 0-5)
    dim_labels.append('Consistency\n(% Score 5)')
    consistency = {}
    for m in models:
        total = len(scores_all[m])
        count_5 = sum(1 for s in scores_all[m] if s == 5)
        consistency[m] = (count_5 / total) * 5 if total > 0 else 0
    dims.append(consistency)

    # 4. Robustness (inverse of std dev, scaled to 0-5)
    dim_labels.append('Robustness\n(Low Variance)')
    robust = {}
    for m in models:
        std = np.std(scores_all[m]) if scores_all[m] else 0
        robust[m] = max(0, 5 - std * 3.33)
    dims.append(robust)

    # 5. Overall Average
    dim_labels.append('Overall\nAverage')
    dims.append({m: np.mean(scores_all[m]) if scores_all[m] else 0 for m in models})

    N = len(dim_labels)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor(BG_COLOR)
    ax.set_facecolor(BG_COLOR)

    for model in models:
        values = [dims[i][model] for i in range(N)]
        values += values[:1]
        ax.plot(angles, values, 'o-', linewidth=2.2, label=model,
                color=COLORS[model], markersize=6, alpha=0.9)
        ax.fill(angles, values, alpha=0.10, color=COLORS[model])

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(dim_labels, fontsize=10, fontweight='medium',
                       ha='center', color='#333333')

    ax.set_ylim(0, 5.5)
    ax.set_yticks([1, 2, 3, 4, 5])
    ax.set_yticklabels(['1', '2', '3', '4', '5'], fontsize=8.5, color='#888888')
    ax.yaxis.grid(True, color=GRID_COLOR, linewidth=0.6)
    ax.xaxis.grid(True, color=GRID_COLOR, linewidth=0.6)

    # Data value annotations — offset per model to avoid overlap
    model_offsets = {
        'GPT':      (-22, -12),
        'DeepSeek': (0, 12),
        'Claude':   (18, -4),
    }
    for model in models:
        values = [dims[i][model] for i in range(N)]
        ox, oy = model_offsets[model]
        for angle, val in zip(angles[:-1], values):
            ax.annotate(
                f'{val:.2f}',
                xy=(angle, val),
                xytext=(ox, oy),
                textcoords='offset points',
                fontsize=8, fontweight='bold',
                color=COLORS[model], alpha=0.85
            )

    ax.legend(
        loc='upper right', bbox_to_anchor=(1.28, 1.12),
        frameon=True, framealpha=0.95, edgecolor='#DDDDDD',
        fancybox=True, fontsize=10.5
    )

    ax.set_title(
        'Model Capability Spectrum\n(Multi-Dimensional Performance Profile)',
        fontweight='bold', fontsize=14, pad=28, color='#1a1a1a'
    )

    plt.tight_layout()
    out = f'{OUTPUT_DIR}/chart3_radar.png'
    fig.savefig(out, bbox_inches='tight')
    print(f'✅ Saved: {out}')
    plt.close()


# Generate All Charts
if __name__ == '__main__':
    create_chart1()
    create_chart2()
    create_chart3()
    print('\n🎉 All three charts generated successfully!')
